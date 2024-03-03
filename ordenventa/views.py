from rest_framework import generics
from django.http import HttpResponse, JsonResponse
from inventario.models import Inventario
from reportes.models import Proveedor
from .models import OrdenDeCompra, OrdenVenta, ItemOrdenVenta
from .serializers import OrdenVentaSerializer, ItemOrdenVentaSerializer
from django.shortcuts import redirect, render, get_object_or_404
from django.views import View
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, JsonResponse
from .forms import ItemOrdenVentaForm, OrdenVentaForm, OrdenVentaUploadForm
from openpyxl import load_workbook
from django.db import transaction
from django.template.loader import render_to_string
from django.utils.dateparse import parse_date


# antiguos view debemos depurar despues
class OrdenVentaCRUDView(View):
    template_name = "ordenventa_crud.html"

    def get(self, request, *args, **kwargs):
        # Obtén la acción de la URL (edit o delete)
        action = request.GET.get("action", None)

        # Lógica para editar
        if action == "edit":
            orden_venta_id = request.GET.get("edit", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                form = OrdenVentaForm(instance=orden_venta)
            else:
                form = OrdenVentaForm()

        # Lógica para eliminar
        elif action == "delete":
            orden_venta_id = request.GET.get("delete", None)
            if orden_venta_id:
                orden_venta = get_object_or_404(OrdenVenta, pk=orden_venta_id)
                orden_venta.delete()
                return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))

        # Lógica por defecto para mostrar la lista y el formulario de creación
        else:
            form = OrdenVentaForm()

        ordenes_venta = OrdenVenta.objects.all().order_by("-id")
        return render(
            request, self.template_name, {"ordenes_venta": ordenes_venta, "form": form}
        )

    def post(self, request, *args, **kwargs):
        form = OrdenVentaForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse_lazy("ordenventa-crud"))
        else:
            ordenes_venta = OrdenVenta.objects.all()
            return render(
                request,
                self.template_name,
                {"ordenes_venta": ordenes_venta, "form": form},
            )


# FUNCION PARA PROCESAR EL EXCEL Y ITEMS INDIVIDUALES
def procesar_orden_venta_excel(request, ordenventa_id):
    if request.method == "POST":
        if (
            "submit-excel" in request.POST
        ):  # Identificar si se envió el formulario de Excel
            form_excel = OrdenVentaUploadForm(request.POST, request.FILES)
            if form_excel.is_valid():
                archivo_excel = request.FILES["archivo_excel"]
                workbook = load_workbook(archivo_excel)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is None or row[1] is None or row[2] is None:
                        continue
                    ItemOrdenVenta.objects.create(
                        ordenventa_id=ordenventa_id,
                        nro_articulo=row[0],
                        desc_articulo=row[1],
                        cantidad=row[2],
                        precio_bruto=row[6],
                        total_bruto=row[12],
                    )
                return redirect("ver_items_orden_venta", ordenventa_id=ordenventa_id)
            else:
                form_item = (
                    ItemOrdenVentaForm()
                )  # Inicializar formulario de ítem vacío si no se procesa el de Excel
        elif (
            "submit-item" in request.POST
        ):  # Identificar si se envió el formulario de ítem individual
            form_item = ItemOrdenVentaForm(request.POST)
            form_excel = (
                OrdenVentaUploadForm()
            )  # Inicializar formulario de Excel vacío por defecto
            if form_item.is_valid():
                nuevo_item = form_item.save(commit=False)
                nuevo_item.ordenventa_id = (
                    ordenventa_id  # Asignar manualmente el ID de la orden de venta
                )
                nuevo_item.save()
                return redirect("ver_items_orden_venta", ordenventa_id=ordenventa_id)
    else:
        form_excel = OrdenVentaUploadForm()
        form_item = ItemOrdenVentaForm()

    return render(
        request,
        "formulario_orden_venta.html",
        {
            "form_excel": form_excel,
            "form_item": form_item,
            "ordenventa_id": ordenventa_id,
        },
    )


# VER LOS ITEMS SEGUN ID
def ver_items_orden_venta(request, ordenventa_id):
    ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    items = ItemOrdenVenta.objects.filter(ordenventa=ordenventa)

    # Verificar si la solicitud es una solicitud HTMX
    if request.headers.get("HX-Request", "false").lower() == "true":
        html = render_to_string("fragmentos/items_orden_venta.html", {"items": items})
        return JsonResponse({"html": html})

    return render(
        request,
        "ver_items_orden_venta.html",
        {"ordenventa": ordenventa, "items": items},
    )


# ORDEN AUTOMATICAS
def ver_ordenes_compra(request, ordenventa_id):
    ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
    ordenes_compra = OrdenDeCompra.objects.filter(
        item_orden_venta__ordenventa=ordenventa,
        # Aquí puedes descomentar y agregar cualquier otro filtro que necesites.
        # item_orden_venta__enviado=True
    )
    # Aquí agregas la obtención de todos los proveedores disponibles.
    proveedores = Proveedor.objects.all()

    return render(
        request,
        "ver_orden_compra.html",
        {
            "ordenes_compra": ordenes_compra,
            "ordenventa": ordenventa,
            "proveedores": proveedores,  # Pasas la lista de proveedores a la plantilla.
        },
    )


# mas ordenes automaticas :P (y es la mas importante pelotudo no la friegues esta vez)
def actualizar_orden_de_compra(request, id):
    if request.method == "POST":
        orden_de_compra = get_object_or_404(OrdenDeCompra, pk=id)

        # Campos existentes
        cantidad = request.POST.get("cantidad", "")
        precio_actual = request.POST.get("precio_actual", "")
        igv = request.POST.get("igv", "")
        detraccion = request.POST.get("detraccion", "")

        # Nuevos campos
        clase = request.POST.get("clase", "")
        proveedor_id = request.POST.get(
            "proveedor", ""
        )  # Cambiado a proveedor_id para claridad
        cuotas = request.POST.get("cuotas", "")
        fecha_pago = request.POST.get(
            "fecha_pago", ""
        )  # Nuevo campo para fecha de pago
        comprobante_pago = request.FILES.get(
            "comprobante_pago", None
        )  # Nuevo campo para el PDF

        # Asegurarse de convertir a tipos numéricos antes de realizar cálculos
        try:
            orden_de_compra.cantidad = (
                int(cantidad) if cantidad else orden_de_compra.cantidad
            )
            orden_de_compra.precio_actual = (
                float(precio_actual) if precio_actual else orden_de_compra.precio_actual
            )
            orden_de_compra.igv = float(igv) if igv else orden_de_compra.igv
            orden_de_compra.detraccion = (
                float(detraccion) if detraccion else orden_de_compra.detraccion
            )
            orden_de_compra.clase = clase if clase else orden_de_compra.clase
            orden_de_compra.cuotas = int(cuotas) if cuotas else orden_de_compra.cuotas

            # Actualización del proveedor
            if proveedor_id:
                proveedor = Proveedor.objects.filter(id=proveedor_id).first()
                orden_de_compra.proveedor = proveedor
            else:
                orden_de_compra.proveedor = None

            # Actualizar fecha_pago si se proporciona
            if fecha_pago:
                orden_de_compra.fecha_pago = parse_date(fecha_pago)

            # Actualizar comprobante_pago si se proporciona
            if comprobante_pago:
                orden_de_compra.comprobante_pago = comprobante_pago

            orden_de_compra.save()
            # Aquí puedes decidir cómo responder después de actualizar la orden de compra.
            return HttpResponse(
                f"Orden de compra actualizada con éxito: {orden_de_compra.precio_total}"
            )

        except ValueError:
            # Manejo de error o devolver un mensaje al usuario
            return HttpResponse("Error en los datos proporcionados", status=400)

    else:
        # Método no permitido
        return HttpResponse("Método no permitido", status=405)


# ordenes d epago :) gracias totales


# experimento :=
def ver_ordenes_pago(request, ordenventa_id=None):
    ordenes_de_pago = []
    ordenventa = None
    fecha = None  # Inicializa fecha aquí

    if ordenventa_id:
        ordenventa = get_object_or_404(OrdenVenta, pk=ordenventa_id)
        ordenes_de_pago = OrdenDeCompra.objects.filter(
            item_orden_venta__ordenventa=ordenventa, precio_total__gt=0.00
        ).select_related("proveedor")
    else:
        fecha_str = request.GET.get("fecha_pago")
        if fecha_str:
            fecha = parse_date(fecha_str)
            if fecha:
                ordenes_de_pago = OrdenDeCompra.objects.filter(
                    fecha_pago=fecha, precio_total__gt=0.00
                ).select_related("proveedor")

    context = {
        "ordenes_de_pago": ordenes_de_pago,
        "ordenventa": ordenventa,
        "fecha": fecha,  # Usa la variable fecha directamente
    }
    return render(request, "ver_orden_pago.html", context)


# desde aqui no hay nada interesante :P


# Vistas para OrdenVenta
class ListaOrdenesVenta(generics.ListCreateAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


class DetalleOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = OrdenVenta.objects.all()
    serializer_class = OrdenVentaSerializer


# Vistas para ItemOrdenVenta
class ListaItemsOrdenVenta(generics.ListCreateAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer


class DetalleItemOrdenVenta(generics.RetrieveUpdateDestroyAPIView):
    queryset = ItemOrdenVenta.objects.all()
    serializer_class = ItemOrdenVentaSerializer
